"""
Importadores
============
Dos importadores complementarios:

  importar_referencia(fuentes, ...)         -- fuente de referencia (BBDD, aforos, HCALL).
  importar_programacion_v2(ruta, ...)  -- Nuevo libro con la programacion
                                          tabular y el modelo operacional por
                                          cruce. Es la FUENTE DE VERDAD para
                                          versiones, planes y fases.

Las hojas se identifican por nombre y por etiqueta de texto, no por filas
fijas. No hay listas de cruces hardcodeadas.
"""
from __future__ import annotations
import csv
import datetime
import sqlite3
import unicodedata
from pathlib import Path

import openpyxl

from . import config as C
from .config import FuenteReferencia


# ----------------------------- utilidades --------------------------- #
def to_seconds(v):
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.time)):
        return v.hour * 3600 + v.minute * 60 + v.second
    if isinstance(v, (int, float)):
        return int(round(v * 86400)) if 0 <= v < 2 else int(round(v))
    s = str(v).strip()
    if s in ('', '#DIV/0!', '#NAME?', '#VALUE!'):
        return None
    if ':' in s:
        p = [float(x) for x in s.split(':')] + [0, 0, 0]
        return int(round(p[0] * 3600 + p[1] * 60 + p[2]))
    try:
        return int(round(float(s.replace(',', '.'))))
    except ValueError:
        return None


def to_float_cl(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(',', '.'))
    except ValueError:
        return None


def norm(s):
    if s is None:
        return ''
    s = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode()
    return ' '.join(s.lower().split())


# Alias de nombres entre fuentes (la BBDD usa "Conavicop", el archivo nuevo
# usa "Conavicop" o "Conavicoop"; Escuadron 2 vs Parque Escuadron 2, etc.)
ALIAS_CRUCE = {
    'conavicoop':         'conavicop',
    'parque escuadron 2': 'escuadron 2',
    'heroes de la concepcion': 'heroes de la concepcion',
}


def _crear_db(ruta: Path, schema: Path) -> sqlite3.Connection:
    if ruta.exists():
        ruta.unlink()
    con = sqlite3.connect(ruta)
    con.executescript(schema.read_text(encoding='utf-8'))
    con.execute('PRAGMA foreign_keys = ON')
    return con


# ------------------- deteccion de bloques HCALL --------------------- #
def _fila_marca(ws, texto: str) -> int | None:
    for r in range(1, ws.max_row + 1):
        if norm(ws.cell(row=r, column=1).value).startswith(norm(texto)):
            return r
    return None


def _leer_bloque_hcall(ws, fila_marca: int) -> dict[str, list[int]]:
    out: dict[str, list[int]] = {}
    r = fila_marca + 1
    if norm(ws.cell(row=r, column=1).value).startswith(norm(C.HCALL_FILA_CRUCE_ID)):
        r += 1
    while r <= ws.max_row:
        etiqueta = ws.cell(row=r, column=1).value
        n = norm(etiqueta)
        if n == '' or any(n.startswith(norm(m)) for m in C.HCALL_MARCA_FIN) \
                or n.startswith(norm(C.HCALL_MARCA_OUT)):
            break
        tiempos = sorted(
            t for t in (to_seconds(ws.cell(row=r, column=c).value)
                        for c in range(2, ws.max_column + 1)) if t is not None)
        if tiempos:
            out[str(etiqueta).strip()] = tiempos
        r += 1
    return out


# ------------------------- carga por base --------------------------- #
def _cargar_infraestructura_base(wb_orig, dir_data: Path, dir_schema: Path):
    """Carga estaciones, cruces, barrera, tramo desde el fuente de referencia.

    No carga programaciones ni planes (esos vienen del archivo v2).
    Devuelve cruce_id_map: {nombre_norm -> cruce_id}.
    """
    con = _crear_db(dir_data / 'infraestructura.db',
                    dir_schema / '1_infraestructura.sql')

    # estaciones desde Itinerario + BBDD
    it = wb_orig[C.HOJAS['itinerario']]
    estaciones, vistas = [], set()
    for r in range(12, it.max_row + 1):
        nom = it.cell(row=r, column=1).value
        if nom and norm(nom) not in vistas:
            vistas.add(norm(nom)); estaciones.append(str(nom).strip())

    bbdd = wb_orig[C.HOJAS['bbdd']]
    cb = C.COLS_BBDD
    for r in range(2, bbdd.max_row + 1):
        if norm(bbdd.cell(row=r, column=cb['nombre']).value) == 'observaciones':
            break
        for col in (cb['estacion_cercana'], cb['tramo_cw_desde'],
                    cb['tramo_cw_hasta'], cb['tramo_cc_desde'], cb['tramo_cc_hasta']):
            nom = bbdd.cell(row=r, column=col).value
            if nom and norm(nom) not in vistas:
                vistas.add(norm(nom)); estaciones.append(str(nom).strip())

    est_id = {}
    for i, nom in enumerate(estaciones, 1):
        con.execute('INSERT INTO estaciones (estacion_id,nombre,orden_linea) '
                    'VALUES (?,?,?)', (i, nom, i))
        est_id[norm(nom)] = i

    cruce_id = {}
    alarma = to_seconds(bbdd.cell(row=C.BBDD_ALARMA[0], column=C.BBDD_ALARMA[1]).value)
    for r in range(2, bbdd.max_row + 1):
        cid = bbdd.cell(row=r, column=cb['id']).value
        nombre = bbdd.cell(row=r, column=cb['nombre']).value
        if cid is None or norm(nombre) == 'observaciones':
            break
        sent = bbdd.cell(row=r, column=cb['sentido_afectacion']).value
        con.execute(
            'INSERT INTO cruces (cruce_id,nombre,comuna,latitud,longitud,'
            'num_pistas_total,num_carriles_lateral,tiene_semaforo,'
            'afecta_lateral,sentido_afectacion,estacion_cercana_id,'
            'dist_estacion_m,estado_camaras) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)',
            (int(cid), str(nombre).strip(),
             bbdd.cell(row=r, column=cb['comuna']).value,
             to_float_cl(bbdd.cell(row=r, column=cb['latitud']).value),
             to_float_cl(bbdd.cell(row=r, column=cb['longitud']).value),
             bbdd.cell(row=r, column=cb['pistas']).value, 2,
             1 if norm(bbdd.cell(row=r, column=cb['semaforo']).value) == 'si' else 0,
             1 if norm(bbdd.cell(row=r, column=cb['afecta_lateral']).value) == 'si' else 0,
             sent if sent in ('CC', 'CW') else None,
             est_id.get(norm(bbdd.cell(row=r, column=cb['estacion_cercana']).value)),
             to_float_cl(bbdd.cell(row=r, column=cb['dist_estacion']).value),
             bbdd.cell(row=r, column=cb['estado_camaras']).value))
        cruce_id[norm(nombre)] = int(cid)
        for sentido, col in (('CW', cb['barrera_cw']), ('CC', cb['barrera_cc'])):
            t = to_seconds(bbdd.cell(row=r, column=col).value)
            if t is not None:
                con.execute('INSERT INTO parametros_barrera (cruce_id,sentido,'
                            'tiempo_barrera_s,margen_pre_s,margen_post_s,'
                            'tiempo_alarma_s,fuente) VALUES (?,?,?,?,?,?,?)',
                            (int(cid), sentido, t, 10, 10, alarma, 'Registro de referencia'))
        dist_tot = to_float_cl(bbdd.cell(row=r, column=cb['dist_total']).value)
        dist_des = to_float_cl(bbdd.cell(row=r, column=cb['dist_estacion']).value)
        for sentido, c1, c2 in (('CW', cb['tramo_cw_desde'], cb['tramo_cw_hasta']),
                                ('CC', cb['tramo_cc_desde'], cb['tramo_cc_hasta'])):
            con.execute('INSERT INTO cruce_tramo (cruce_id,sentido,'
                        'estacion_desde_id,estacion_hasta_id,dist_desde_m,'
                        'dist_total_m) VALUES (?,?,?,?,?,?)',
                        (int(cid), sentido,
                         est_id.get(norm(bbdd.cell(row=r, column=c1).value)),
                         est_id.get(norm(bbdd.cell(row=r, column=c2).value)),
                         dist_des, dist_tot))
    con.commit()
    res = {t: con.execute(f'SELECT count(*) FROM {t}').fetchone()[0]
           for t in ('estaciones', 'cruces')}
    con.close()
    return cruce_id, res


def _cargar_demanda(wbs, fuentes, cruce_id, dir_data: Path, dir_schema: Path):
    """Carga aforos (Laboral) + HCALL + itinerario desde fuente de referencia."""
    con = _crear_db(dir_data / 'demanda.db', dir_schema / '2_demanda.sql')

    for cid, f in enumerate(fuentes, 1):
        wb = wbs[f.ruta]
        con.execute('INSERT INTO campanias_medicion (campania_id,nombre,'
                    'descripcion) VALUES (?,?,?)',
                    (cid, f.campania, 'Importado de la fuente de referencia ' + Path(f.ruta).name))
        lg = wb[C.HOJAS['llegadas']]; cl = C.COLS_LLEGADAS
        for r in range(2, lg.max_row + 1):
            cruce = lg.cell(row=r, column=cl['cruce']).value
            kid = cruce_id.get(norm(cruce))
            t_ini = to_seconds(lg.cell(row=r, column=cl['t_ini']).value)
            t_fin = to_seconds(lg.cell(row=r, column=cl['t_fin']).value)
            veh_h = lg.cell(row=r, column=cl['veh_h']).value
            if kid is None or t_ini is None or veh_h is None:
                continue
            # Los aforos del fuente de referencia son de dia LABORAL.
            con.execute('INSERT OR IGNORE INTO llegadas_vehiculares (campania_id,'
                        'cruce_id,tipo_dia,t_inicio_s,t_fin_s,flujo_veh_h) '
                        'VALUES (?,?,?,?,?,?)',
                        (cid, kid, 'Laboral', t_ini, min(t_fin, 86400),
                         float(veh_h)))

    con.execute('INSERT INTO itinerario_versiones (itinerario_id,nombre,'
                'descripcion) VALUES (1,?,?)',
                ('Itinerario L2 base', 'Malla operacional de referencia'))
    hc = wbs[fuentes[0].ruta][C.HOJAS['hcall']]
    fila_in = _fila_marca(hc, C.HCALL_MARCA_IN)
    fila_out = _fila_marca(hc, C.HCALL_MARCA_OUT)
    ins = _leer_bloque_hcall(hc, fila_in) if fila_in else {}
    outs = _leer_bloque_hcall(hc, fila_out) if fila_out else {}
    n_ev = 0
    for nombre, lin in ins.items():
        kid = cruce_id.get(norm(nombre))
        lout = outs.get(nombre)
        if kid is None or not lout:
            continue
        for i in range(min(len(lin), len(lout))):
            con.execute('INSERT INTO eventos_barrera (itinerario_id,cruce_id,'
                        'instante_paso_s,hcall_in_s,hcall_out_s) '
                        'VALUES (1,?,?,?,?)', (kid, lout[i] - 10, lin[i], lout[i]))
            n_ev += 1
    con.commit()
    res = {'llegadas': con.execute('SELECT count(*) FROM llegadas_vehiculares').fetchone()[0],
           'hcall': n_ev,
           'cruces_hcall': len({norm(k) for k in ins} & {norm(k) for k in cruce_id})}
    con.close()
    return res


def _cargar_escenarios_vacios(dir_data: Path, dir_schema: Path):
    con = _crear_db(dir_data / 'escenarios.db', dir_schema / '3_escenarios.sql')
    con.close()


def importar_referencia(fuentes: list[FuenteReferencia], dir_data: Path,
                   dir_schema: Path, verbose: bool = True) -> dict:
    """Importa el fuente de referencia: BBDD, aforos, HCALL, itinerario.

    NO importa programaciones/planes (esos vienen del archivo v2).
    """
    wbs = {f.ruta: openpyxl.load_workbook(f.ruta, data_only=True) for f in fuentes}
    cruce_id, r_inf = _cargar_infraestructura_base(wbs[fuentes[0].ruta],
                                                    dir_data, dir_schema)
    r_dem = _cargar_demanda(wbs, fuentes, cruce_id, dir_data, dir_schema)
    _cargar_escenarios_vacios(dir_data, dir_schema)
    for wb in wbs.values():
        wb.close()
    resumen = {**r_inf, **r_dem}
    if verbose:
        print(f"  infraestructura.db : {resumen['estaciones']} estaciones, "
              f"{resumen['cruces']} cruces (sin programacion)")
        print(f"  demanda.db         : {resumen['llegadas']} bandas de aforo, "
              f"{resumen['hcall']} eventos HCALL en {resumen['cruces_hcall']} cruces")
    return resumen


# ==================================================================== #
#  IMPORTADOR DEL ARCHIVO v2 (programaciones, planes, modelo operacional)
# ==================================================================== #
def importar_programacion_v2(ruta_xlsx: Path, dir_data: Path,
                              verbose: bool = True) -> dict:
    """Carga el archivo de programacion v2 a infraestructura.db.

    Pobla:
      - versiones_programacion        (hoja versiones_programacion)
      - planes_horarios_cruce         (hoja planes_horarios_cruce)
      - programacion_fases            (hoja programacion_fases_revision)
      - modelo_operacional_cruce      (hoja modelo_operacional_cruce)
      - cruces_reconfiguracion        (hoja asignacion_cruce_programa)
    """
    wb = openpyxl.load_workbook(ruta_xlsx, data_only=True)
    con = sqlite3.connect(dir_data / 'infraestructura.db')
    con.row_factory = sqlite3.Row
    con.execute('PRAGMA foreign_keys = ON')

    cruce_id = {norm(r['nombre']): r['cruce_id']
                for r in con.execute('SELECT cruce_id, nombre FROM cruces')}

    def resolver_cruce(nom):
        clave = ALIAS_CRUCE.get(norm(nom), norm(nom))
        return cruce_id.get(clave)

    # --- versiones_programacion ---
    ws = wb['versiones_programacion']
    hdr = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(hdr, row))
        if d.get('version_prog_id') is None:
            continue
        con.execute(
            'INSERT OR REPLACE INTO versiones_programacion (version_prog_id,'
            'nombre,fecha,tipo_version,fuente,descripcion) VALUES (?,?,?,?,?,?)',
            (int(d['version_prog_id']), d['nombre'], str(d.get('fecha') or ''),
             d.get('tipo_version'), d.get('fuente'), d.get('descripcion')))

    # --- planes_horarios_cruce ---
    ws = wb['planes_horarios_cruce']
    hdr = [c.value for c in ws[1]]
    n_planes, no_match = 0, set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(hdr, row))
        kid = resolver_cruce(d['cruce'])
        if kid is None:
            no_match.add(d['cruce']); continue
        ini = to_seconds(d['hora_inicio']); fin = to_seconds(d['hora_fin'])
        if ini is None or fin is None:
            continue
        # 24:00:00 (fin de dia) puede venir como 0; tratarlo como 86400.
        if fin == 0 and ini > 0:
            fin = 86400
        con.execute(
            'INSERT OR REPLACE INTO planes_horarios_cruce (version_prog_id,'
            'cruce_id,tipo_dia,hora_inicio_s,hora_fin_s,plan_id,fuente) '
            'VALUES (?,?,?,?,?,?,?)',
            (int(d['version_prog_id']), kid, d['tipo_dia'], ini, fin,
             int(d['plan_id']), d.get('fuente')))
        n_planes += 1

    # --- programacion_fases (desde la hoja _revision) ---
    ws = wb['programacion_fases_revision']
    hdr = [c.value for c in ws[1]]
    con.execute('DELETE FROM programacion_fases')
    n_fases, n_filtradas = 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(hdr, row))
        kid = resolver_cruce(d['cruce'])
        if kid is None:
            no_match.add(d['cruce']); continue
        try:
            dur = int(d['duracion_s'])
        except (TypeError, ValueError):
            continue
        if dur <= 0:
            # Fase fantasma (artefacto del proceso de rotacion del archivo
            # fuente). Saltada para no romper el motor.
            n_filtradas += 1
            continue
        try:
            con.execute(
                'INSERT INTO programacion_fases (version_prog_id,cruce_id,'
                'plan_id,fase_id,duracion_s,entreverde_s,cum_inicio_s,'
                'cum_fin_s,es_verde_lateral,ciclo_s,fase_origen,fuente,'
                'confianza,estado_carga) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
                (int(d['version_prog_id']), kid, int(d['plan_id']),
                 int(d['fase_id']), dur,
                 int(d['entreverde_s'] or 0), int(d['cum_inicio_s']),
                 int(d['cum_fin_s']), int(d['es_verde_lateral'] or 0),
                 int(d['ciclo_s']), d.get('fase_origen'), d.get('fuente'),
                 d.get('confianza_asignacion') or d.get('confianza'),
                 d.get('estado_carga')))
            n_fases += 1
        except sqlite3.IntegrityError:
            pass

    # --- modelo_operacional_cruce ---
    ws = wb['modelo_operacional_cruce']
    hdr = [c.value for c in ws[1]]
    n_modelo = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(hdr, row))
        kid = resolver_cruce(d['cruce'])
        if kid is None:
            continue
        con.execute(
            'INSERT OR REPLACE INTO modelo_operacional_cruce (cruce_id,'
            'tipo_modelo,usa_reconfiguracion,version_prog_id,descripcion,'
            'fuente) VALUES (?,?,?,?,?,?)',
            (kid, d['tipo_modelo'], int(d['usa_reconfiguracion']),
             int(d['version_prog_id']), d.get('descripcion'), d.get('fuente')))
        n_modelo += 1

    # --- cruces_reconfiguracion (desde asignacion_cruce_programa) ---
    ws = wb['asignacion_cruce_programa']
    hdr = [c.value for c in ws[1]]
    con.execute('DELETE FROM cruces_reconfiguracion')
    n_asig = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(hdr, row))
        if d.get('modelo_operacional') != 'RECONFIG':
            continue
        kid = resolver_cruce(d['cruce_ferroviario'])
        if kid is None:
            continue
        con.execute(
            'INSERT OR REPLACE INTO cruces_reconfiguracion (cruce_id,'
            'via_principal,codigo_proyecto,comuna_referencia,estado_carga,'
            'confianza,fuente) VALUES (?,?,?,?,?,?,?)',
            (kid, d.get('programacion_fuente_elegida') or d.get('programacion_fuente'),
             str(d.get('id_interseccion') or ''),
             None, d.get('estado_carga'), d.get('confianza'),
             d.get('archivo_fuente')))
        n_asig += 1

    con.commit(); con.close(); wb.close()
    if verbose:
        print(f"  programacion v2    : {n_planes} planes, {n_fases} fases, "
              f"{n_modelo} cruces con modelo operacional")
        if n_filtradas:
            print(f"                       {n_filtradas} fases fantasma "
                  f"(dur<=0) filtradas")
        print(f"                       {n_asig} cruces declarados RECONFIG")
        if no_match:
            print(f"  WARN: no emparejados con BBDD: {sorted(no_match)}")
    return {'planes': n_planes, 'fases': n_fases,
            'fases_filtradas': n_filtradas,
            'modelo_operacional': n_modelo, 'reconfig': n_asig}
