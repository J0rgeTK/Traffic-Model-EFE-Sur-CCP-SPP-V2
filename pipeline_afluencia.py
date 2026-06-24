"""
Pipeline afluencia EFE/Fesur -> proyeccion mensual 2027
=======================================================
Semilla de repo. Dos etapas:
  1. ETL  : extrae afluencia DIARIA por servicio desde los 4 formatos heterogeneos.
  2. FORECAST: descomposicion clasica (indice estacional * nivel * tendencia amortiguada).

El ajuste climatico NO se modela aqui: se mide aparte a nivel diario (ver
`efecto_lluvia_diario`) y se aplica como factor sobre el agregado mensual.

Requisitos: pandas, numpy, openpyxl.  (matplotlib opcional para graficar)
Autor: prototipo v0. Revisar CAVEATS antes de usar en presupuesto.
"""
import re, glob, os
import numpy as np
import pandas as pd
import openpyxl

# --------------------------------------------------------------------------
# CAVEATS CONOCIDOS (v0) -- no ignorar:
#  - BIOTREN: nivel ETL ~5-13% bajo el Resumen oficial (definicion de "Pasajeros"
#             difiere: excluye Salidas sin validar / compensaciones). Reconciliar
#             con el equipo que arma el Resumen antes de usar montos absolutos.
#             Ademas ~20% de dias faltantes.
#  - LLANQUIHUE_PM (=XP-NQ en RROO): solo 13 meses, 28% dias faltantes, salto
#             ene-feb 2026 (~2.5x). Confianza BAJA. Confirmar si el salto es
#             demanda real (peak verano) o cambio de definicion ANTES de proyectar.
#  - Codigos de estacion RROO (XP, NQ, AL, EV) != nombres presupuesto
#             (Llanquihue, Puerto Varas, Alerce, La Paloma): falta diccionario.
# --------------------------------------------------------------------------

SVC_CODE = re.compile(r'^\d{4,6}$')      # cabecera = codigo de tren (columna de servicio)
BT_STOP  = {'salidas', 'total', 'afluencias', 'carga', 'dif', '%', 'pm', 'pt', 'pp'}

# Tipos de dia que se consideran esperados para normalizar meses incompletos.
# Evita sobreimputar servicios que no operan fines de semana, especialmente
# Llanquihue-Puerto Montt, cuyo itinerario informado es solo lunes-viernes.
EXPECTED_DT_BY_SERVICE = {
    'BIOTREN': ['LV', 'Sab', 'Dom'],
    'CORTO_LAJA': ['LV', 'Sab', 'Dom'],
    'TREN_ARAUCANIA': ['LV', 'Sab', 'Dom'],
    'LLANQUIHUE_PM': ['LV'],
}



def _num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_matrix(path, sheet, servicio):
    """CL / LP / TA: hoja matriz Fecha x codigo_tren. Suma SOLO columnas de servicio
    (excluye 'Total general', 'Tipo dia', 'Recaudacion' y la fila Total al pie)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        return []
    rows = list(wb[sheet].iter_rows(values_only=True))
    hdr = rows[0]
    cols = [i for i, h in enumerate(hdr) if h is not None and SVC_CODE.match(str(h).strip())]
    out = []
    for r in rows[1:]:
        if not r or r[0] is None:
            continue
        f = pd.to_datetime(r[0], dayfirst=True, errors='coerce')
        if pd.isna(f) or f.year < 2020:          # descarta fila Total / basura
            continue
        out.append((servicio, f.normalize(), sum(_num(r[i]) for i in cols if i < len(r))))
    return out


def _parse_biotren(path):
    """BIOTREN: lee la hoja diaria.

    Criterio de lectura:
    - Si existe la columna "Afluencias + Multas +SSE", se usa como total oficial
      diario.
    - Si no existe, se usa como respaldo la suma del bloque de tipos de
      pasajero ubicado antes de las columnas resumen.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = [s for s in wb.sheetnames if 'diaria' in s.lower()]
    if not sh:
        return []
    rows = list(wb[sh[0]].iter_rows(values_only=True))
    h = next((i for i, r in enumerate(rows)
              if r and len(r) > 1 and str(r[1]).strip().upper() == 'FECHA'), None)
    if h is None:
        return []
    hdr = rows[h]

    total_col = None
    for preferida in ['Afluencias + Multas +SSE', 'Afluencias + Multas ', 'TOTAL AFLUENCIA ']:
        for i, v in enumerate(hdr):
            if isinstance(v, str) and v.strip().lower() == preferida.strip().lower():
                total_col = i
                break
        if total_col is not None:
            break

    end = len(hdr)
    if total_col is None:
        for i in range(2, len(hdr)):
            hl = str(hdr[i]).strip().lower() if hdr[i] is not None else ''
            if any(hl.startswith(s) for s in BT_STOP):
                end = i
                break

    out = []
    for r in rows[h + 1:]:
        if not r or r[1] is None:
            continue
        f = pd.to_datetime(r[1], dayfirst=True, errors='coerce')
        if pd.isna(f) or f.year < 2020:
            continue
        if total_col is not None:
            pax = _num(r[total_col]) if total_col < len(r) else 0.0
        else:
            pax = sum(_num(r[i]) for i in range(2, end) if i < len(r))
        out.append(("BIOTREN", f.normalize(), pax))
    return out

def etl_afluencia_diaria(base_bbdd):
    """Recorre BBDD/{BT,CL,LP,TA} y devuelve df tidy: servicio, fecha, pasajeros."""
    rec = []
    for fp in glob.glob(f"{base_bbdd}/BT/*.xlsx"):
        rec += _parse_biotren(fp)
    for fp in glob.glob(f"{base_bbdd}/CL/*.xlsx"):
        wb = openpyxl.load_workbook(fp, read_only=True)
        sh = [s for s in wb.sheetnames if s.lower().startswith('afxdia')]
        if sh:
            rec += _parse_matrix(fp, sh[0], "CORTO_LAJA")
    for code, svc in [("LP", "LLANQUIHUE_PM"), ("TA", "TREN_ARAUCANIA")]:
        for fp in glob.glob(f"{base_bbdd}/{code}/*.xlsx"):
            rec += _parse_matrix(fp, 'AfluDiariaxServ', svc)
    df = pd.DataFrame(rec, columns=['servicio', 'fecha', 'pasajeros'])
    df = df[df['pasajeros'] > 0]
    # dedup: archivos solapan fechas -> conserva el mayor (mas completo)
    df = (df.sort_values('pasajeros')
            .drop_duplicates(['servicio', 'fecha'], keep='last')
            .sort_values(['servicio', 'fecha'])
            .reset_index(drop=True))
    return df


def _dt_label_from_date(fecha):
    dow = fecha.dayofweek
    return 'LV' if dow < 5 else ('Sab' if dow == 5 else 'Dom')


def mensualizar(df_diario, cobertura_min=0.5):
    """Agrega a mensual corrigiendo sesgo por dias faltantes.

    Version actualizada:
    - Para servicios con operacion todos los dias se normaliza contra dias calendario.
    - Para servicios solo lunes-viernes, como Llanquihue-Puerto Montt, se normaliza
      contra dias planificados del mes, no contra fines de semana sin oferta.
    - Retorna columnas auxiliares de cobertura para auditar meses incompletos.
    """
    df = df_diario.copy()
    df['fecha'] = pd.to_datetime(df['fecha'])
    df['mes'] = df['fecha'].dt.to_period('M')
    df['dt'] = df['fecha'].apply(_dt_label_from_date)
    rows = []
    for (s, m), g in df.groupby(['servicio', 'mes']):
        tipos_esperados = EXPECTED_DT_BY_SERVICE.get(s, ['LV', 'Sab', 'Dom'])
        fechas_mes = pd.date_range(m.start_time, m.end_time, freq='D')
        labels_mes = ['LV' if x.dayofweek < 5 else ('Sab' if x.dayofweek == 5 else 'Dom') for x in fechas_mes]
        dias_esperados = sum(x in tipos_esperados for x in labels_mes)
        gg = g[g['dt'].isin(tipos_esperados)]
        dias_obs = gg['fecha'].nunique()
        cobertura = dias_obs / max(dias_esperados, 1)
        pax_norm = gg['pasajeros'].mean() * dias_esperados if dias_obs else np.nan
        rows.append({'servicio': s, 'mes': m, 'cobertura': cobertura,
                     'dias_obs': dias_obs, 'dias_esperados': dias_esperados,
                     'pax_norm': pax_norm})
    mdf = pd.DataFrame(rows)
    if not mdf.empty:
        mdf['mes'] = pd.PeriodIndex(mdf['mes'], freq='M')
    return mdf[(mdf['cobertura'] >= cobertura_min) & mdf['pax_norm'].notna()].sort_values(['servicio', 'mes'])

def proyectar_2027(mdf, anio=2027):
    """Descomposicion clasica multiplicativa. Tendencia amortiguada (+-2%/mes) y
    SOLO si hay >=24 meses; con menos, nivel plano (sin extrapolar)."""
    res = {}
    meta = {}
    for s, g in mdf.groupby('servicio'):
        g = g.sort_values('mes').reset_index(drop=True)
        n = len(g)
        g['m'] = g['mes'].dt.month
        seas = g.groupby('m')['pax_norm'].mean()
        seas = seas / seas.mean()                       # indice estacional (media=1)
        g['des'] = g['pax_norm'] / g['m'].map(seas)
        if n >= 24:
            slope = np.polyfit(np.arange(n), np.log(g['des']), 1)[0]
            gth = float(np.clip(np.expm1(slope), -0.02, 0.02))
        else:
            gth = 0.0                                    # series corta: no extrapolar
        level = g['des'].tail(min(12, n)).mean()
        last = g['mes'].max()
        out = {}
        for k in range(12):
            mm = k + 1
            h = (pd.Period(f'{anio}-01', 'M').ordinal - last.ordinal) + k
            out[f'{anio}-{mm:02d}'] = level * seas.get(mm, 1.0) * ((1 + gth) ** h)
        res[s] = out
        meta[s] = {'n_meses': n, 'tendencia_pct_mes': round(gth * 100, 2),
                   'confianza': 'ALTA' if n >= 24 else ('MEDIA' if n >= 18 else 'BAJA')}
    return pd.DataFrame(res).round(0).astype(int), meta


# --------------------------------------------------------------------------
# ETAPA CLIMA (a completar con datos diarios de lluvia por comuna).
# Mapeo sugerido servicio -> comuna/estacion para Open-Meteo (lat/lon):
#   BIOTREN        -> Concepcion / Coronel        (-36.83, -73.05)
#   CORTO_LAJA     -> Los Angeles / Laja          (-37.28, -72.71)
#   TREN_ARAUCANIA -> Temuco / Victoria           (-38.74, -72.59)
#   LLANQUIHUE_PM  -> Puerto Varas / Puerto Montt (-41.32, -72.99)
# --------------------------------------------------------------------------
def efecto_lluvia_diario(df_diario, df_lluvia, umbral_mm=1.0):
    """Mide cuanto cae la afluencia en dias de lluvia vs dias secos, por servicio.
    df_lluvia: columnas [servicio, fecha, precip_mm].
    Devuelve, por servicio, el factor multiplicativo dia_lluvia/dia_seco
    controlando por dia de semana (evita confundir lluvia con fin de semana)."""
    d = df_diario.merge(df_lluvia, on=['servicio', 'fecha'], how='inner')
    if d.empty:
        return pd.DataFrame()
    d['lluvia'] = d['precip_mm'] >= umbral_mm
    d['dow'] = d['fecha'].dt.dayofweek
    res = []
    for s, g in d.groupby('servicio'):
        # promedio por (dia_semana, lluvia) y luego ratio promediado sobre dias
        piv = g.groupby(['dow', 'lluvia'])['pasajeros'].mean().unstack()
        if piv.shape[1] == 2:
            ratio = (piv[True] / piv[False]).mean()
            res.append({'servicio': s, 'factor_lluvia': round(ratio, 3),
                        'reduccion_pct': round((1 - ratio) * 100, 1),
                        'n_dias_lluvia': int(g['lluvia'].sum())})
    return pd.DataFrame(res)


if __name__ == '__main__':
    BASE = os.environ.get('BBDD', 'data/raw_bbdd')
    diario = etl_afluencia_diaria(BASE)
    diario.to_csv('data/afluencia_diaria_consolidada.csv', index=False)
    mdf = mensualizar(diario)
    proy, meta = proyectar_2027(mdf)
    proy.to_csv('outputs/proyeccion_afluencia_2027_base.csv')
    print(proy.to_string())
    for s, m in meta.items():
        print(s, m)
